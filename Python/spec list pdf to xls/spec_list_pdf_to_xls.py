import slate3k
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, NamedStyle, Font
from openpyxl.styles import Alignment

print("please input file name without .pdf:")
file_name = input()
# file_name = '1312_06'

# check file exists
while not os.path.isfile(file_name+'.pdf'):
        print("file doesn't exist, please re-input file name, or type exit for exit:")
        file_name = input()
        if file_name == "exit":
            exit(0)

# open pdf file with prop: read boolean
with open(file_name+'.pdf', 'rb') as f:
    doc = slate3k.PDF(f)
    print(len(doc))
# print(doc)

# open txt file for writing and write text from pdf
with open('1312.txt', 'w', encoding='utf-8') as fg:
    fg.write(str(doc))

# open txt file for reading and read text to txt0
with open('1312.txt', 'r', encoding='utf-8', errors='ignore') as file:
    txt = file.readline()

# delete txt temporary file
os.remove('1312.txt')

# text clearing
txt_pr = str(txt).split('\\n')
txt_cleared = []
for a in txt_pr:
    if a != '':
        txt_cleared.append(a)

# cells style
main_style = NamedStyle(name='main_style')
minor_style = NamedStyle(name='minor_style')

thin = Side(border_style='thin', color='000000')
main_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
minor_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
main_style.fill = PatternFill('solid', fgColor='6fb1bd')
minor_style.fill = PatternFill('solid', fgColor='b3dee7')
main_style.font = Font(size=12, name='Calibri')
minor_style.font = Font(size=12, name='Calibri')

# open workbook, rename title
wb = Workbook()
ws = wb.active
ws.title = file_name+'_spec_list'
ws.sheet_properties.tabColor = "20FF20"
wb.add_named_style(main_style)
wb.add_named_style(minor_style)

# change width of column (convert 2.85points=1mm)
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 130
ws.column_dimensions["C"].width = 60
ws.column_dimensions["D"].width = 35
ws.column_dimensions["E"].width = 45
ws.column_dimensions["F"].width = 20
ws.column_dimensions["G"].width = 20
ws.column_dimensions["H"].width = 25
ws.column_dimensions["I"].width = 40

ws.row_dimensions[1].height = 50


# insert style in hat
j = 1
while j != 10:
    ws.cell(row=1, column=j).style = main_style
    ws.cell(row=1, column=j).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    j += 1

# hat
a1 = ws.cell(row=1, column=1, value='Поз.')
a2 = ws.cell(row=1, column=2, value='Наименование и техническая характеристика')
a3 = ws.cell(row=1, column=3, value='Тип, марка,\nобозначение документа,\nопросного листа')
a4 = ws.cell(row=1, column=4, value='Код продукции')
a5 = ws.cell(row=1, column=5, value='Поставщик')
a6 = ws.cell(row=1, column=6, value='Ед.изме-рения')
a7 = ws.cell(row=1, column=7, value='Кол.')
a8 = ws.cell(row=1, column=8, value='Масса 1 ед.,\nкг')
a9 = ws.cell(row=1, column=9, value='Примечание')

# combine same data
txt_sum_nc = txt_cleared[14:]
txt_sum = txt_sum_nc[:len(txt_sum_nc)-1]
txt_combine = []
for i in txt_sum:
    m = len(i)
    if (i[0] == '@' or i[0] == '&' or i[0] == '$' or i[0] == '^' or
                i[0] == '*' or i[0] == '<' or i[0] == '>' or
                i[m-2:m] == 'k.' or i[0] == ']' or i[0] == '{'):
        txt_combine.append(i)
    else:
        k = len(txt_combine)
        j = txt_combine[k-1]
        del txt_combine[k-1]
        txt_combine.append(j+'\n'+i)

# combine same data 2
t = '~'
txt_combine2 = []
for i in txt_combine:

    if i[0] != t[0]:
        txt_combine2.append(i)
    else:
        k = len(txt_combine2)
        j = txt_combine2[k - 1]
        del txt_combine2[k - 1]
        txt_combine2.append(j + '\n' + i)
    t = i

print(txt_combine2)


# insert style in data strings
def style_data(r, c, v):
    jj = 1
    if v == 1:
        while jj != c+1:
            ws.cell(row=r, column=jj).style = minor_style
            ws.cell(row=r, column=jj).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            jj += 1
    if v == 2:
            ws.cell(row=r, column=c).style = minor_style
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

# put data in excel
row_num = 1
for i in txt_combine2:
    m = len(i)
    if i[0] == '@' or i[0] == '&':
        row_num += 1
        style_data(row_num, 9, 1)
    if i[0] == '@':
        row_num += 1
        style_data(row_num, 9, 1)
        ws.cell(row=int(row_num), column=int(2), value=str(i).replace('@', ''))
        row_num += 1
        style_data(row_num, 9, 1)
    if i[0] == '&':
        ws.cell(row=int(row_num), column=int(1), value=str(i).replace('&', ''))
    if i[0] == '$':
        ws.cell(row=int(row_num), column=int(2), value=str(i).replace('$', ''))
        style_data(row_num, 2, 2)
    if i[0] == '^':
        ws.cell(row=int(row_num), column=int(3), value=str(i).replace('^', ''))
    if i[0] == '*':
        ws.cell(row=int(row_num), column=int(4), value=str(i).replace('*', ''))
    if i[0] == '<':
        ws.cell(row=int(row_num), column=int(5), value=str(i).replace('<', ''))
    if i[0] == '>':
        ws.cell(row=int(row_num), column=int(6), value=str(i).replace('>', ''))
    if i[m-2:m] == 'k.':
        ws.cell(row=int(row_num), column=int(7), value=str(i).replace(' stk.', ''))
    if i[0] == ']':
        i = str(i).replace('000000 кг', '')
        ws.cell(row=int(row_num), column=int(8), value=str(i).replace(']', ''))
    if i[0] == '{':
        ws.cell(row=int(row_num), column=int(9), value=str(i).replace('{', ''))



'''
j = 2
k = 0
m = 0

for i in txt_sum:
    if m % 5 == 0:
        j += 1
        k = 0
    k += 1
    m += 1
    if k == 1:
        ws.cell(row=int(j), column=int(4), value=txt_sum[m-1])
    if k == 2:
        ws.cell(row=int(j), column=int(5), value=txt_sum[m-1])
    if k == 3:
        ws.cell(row=int(j), column=int(2), value=txt_sum[m-1])
    if k == 4:
        ws.cell(row=int(j), column=int(3), value=txt_sum[m-1])
    if k == 5:
        ws.cell(row=int(j), column=int(1), value=txt_sum[m-1])
    ws.cell(row=int(j), column=int(k)).style = minor_style
    print(j)
'''
# save excel file
wb.save(file_name+'_spec_list.xlsx')

#print(len(txt_pr))
#print(txt_sum)



