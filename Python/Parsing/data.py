def input_check():
    i = input('please enter a digit: ')
    if str(i).isdigit():
        return i
    else:
        print('it is not a digit, please re-enter')
        return input_check()


def data_import(i):
    from openpyxl import load_workbook
    import os
    # open file
    wb = load_workbook(os.path.abspath("Terminals.xlsx"), read_only=True)
    # select work sheet
    sheet_terminals = wb['terminals']
    # counting of data cells
    d = sheet_terminals.max_row
    # checking of existing data in cell, return values
    if i <= d:
        return sheet_terminals['G' + str(i)].value
    else:
        return "None"


def data_save(i, j):
    from openpyxl import load_workbook

    wb = load_workbook("Terminals.xlsx")  # open file
    sheet = wb["ROPTerminals"]  # select work sheet
    m = 0
    n = len(j)
    while m < n:
        sheet.cell(row=i, column=m+1).value = j[m]
        m += 1
    wb.save(filename="Terminals.xlsx")


def data_num():
    from openpyxl import load_workbook
    import os

    wb = load_workbook(os.path.abspath("Terminals.xlsx"), read_only=True)  # open file
    sheet_terminals = wb['terminals']  # select work sheet

    i = sheet_terminals.max_row
    return i


def open_last_num():
    import os
    # open last used num
    if os.path.isfile('temp'):
        temp_file = open('temp', 'r')
        last_used_num = temp_file.read()
        temp_file.close()
    # using function data_import
        return int(last_used_num)
    else:
        return int(1)


def save_num_sync(last_used_num):
    temp_file = open('temp', 'w')
    temp_file.write(last_used_num)
    temp_file.close()


def data_collect(i):
    import requests
    # описание заголовка для запроса
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko',
    }
    r = requests.get(i, headers=headers)

    if r.status_code == requests.codes.ok:
        # считываем текстовые данные
        return r.text
    else:
        return '0'


def data_search(r_text, s_text):
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(r_text, "html.parser")
    # считываем данные
    dirty_desc = soup.find_all("div", class_="pxc-prod-detail-txt")
    dirty_short_desc = soup.h1

    # check existing short description, if none - return '0', because at phoenix site it mean no data!
    if dirty_short_desc is None:
        return '0'

    dirty_tech_data = soup.find_all("table", class_="pxc-tbl")
    # len(soup.find_all("table", class_="pxc-tbl")) can count num of tables, and split them
    # test = soup.find_all("table", class_="pxc-tbl")[0]
    soup.clear()
    # выгружаем текст
    desc = str(dirty_desc)
    soup = BeautifulSoup(desc, "html.parser")
    desc = soup.get_text()
    desc = desc.split("\n")
    # выгружаем текст
    short_desc = str(dirty_short_desc)
    soup = BeautifulSoup(short_desc, "html.parser")
    short_desc = soup.get_text()
    # выгружаем текст
    tech_data = str(dirty_tech_data)
    soup = BeautifulSoup(tech_data, "html.parser")
    tech_data = soup.get_text()
    tech_data = tech_data.split("\n")
    tech_data1 = []
    # пересобираем текстовые данные
    for a in tech_data:
        if a != '' and a != '[' and a != ']' and a !=', ':
            tech_data1.append(a)

    # загружаем данные в парсер
    soup = BeautifulSoup(s_text, "html.parser")
    # считываем данные
    dirty_comm_data = soup.find("table", class_="pxc-tbl")
    soup.clear()
    # выгружаем текст
    comm_data = str(dirty_comm_data)
    soup = BeautifulSoup(comm_data, "html.parser")
    comm_data = soup.get_text()
    comm_data = comm_data.split("\n")
    comm_data1 = []
    # пересобираем текстовые данные
    for a in comm_data:
        if a != '':
            comm_data1.append(a)

    return short_desc, desc[1], tech_data1, comm_data1


def image_search(i):
    import requests
    from bs4 import BeautifulSoup

    # описание заголовка для запроса
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko',
    }
    # запрос данных
    im = requests.get('https://www.phoenixcontact.com/online/portal/ru/?uri=pxc-oc-itemdetail:pid=' + i +
                      '&library=ruru&pcck=P-22-03-01-02&tab=1&selectedCategory=ALL', headers=headers)

    # считываем текстовые данные
    im_text = im.text
    # загружаем данные в парсер
    soup = BeautifulSoup(im_text, "html.parser")
    # считываем данные
    dirty_image_ad = soup.find_all("img", class_="pxc-img")
    # выгружаем текст
    image_src = str(dirty_image_ad).split()
    if im.status_code == requests.codes.ok and len(image_src) > 3:
        image_url_proc = image_src[3].lstrip("\"src=").rstrip("\"")
        image_url_small = "https://www.phoenixcontact.com/" + image_url_proc
        image_url_large = image_url_small.replace("small1", "large")
        image_url_large = image_url_large.replace("int_01", "int_04")
        return image_url_large
    print('image not found')
    return ''


def image_download(i, j):
    import requests
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko',
        }

    url = i
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        with open(j+".jpg", 'wb') as f:
            f.write(response.content)


#    url = "https://www.phoenixcontact.com/assets/images_pr/product_photos/large/25176_1000_int_04.jpg"
#    response = requests.get(url, headers=headers)
#    if response.status_code == 200:
#        with open(".sample.jpg", 'wb') as f:
#            f.write(response.content)
