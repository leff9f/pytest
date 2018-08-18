from multiprocessing import Pool
import requests
import time
from bs4 import BeautifulSoup

headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Trident/7.0; rv:11.0) like Gecko',
        }

toc = [time.time()]


def f(x):
    r = requests.get(
        'https://www.phoenixcontact.com/online/portal/ru/?uri=pxc-oc-itemdetail:pid=' + str(x) + '&library=ruru&pcck=P-22-03-01-02&tab=1&selectedCategory=ALL',
        headers=headers)
    return r.status_code


def mp():
    if __name__ == '__main__':
        p = Pool(7)
        return print(p.map(f, [1411272, 1619445, 1619868, 1619596, '0624047', 1619445, 1619868, 1619596, 1619868, 1619596,
                        1411272, 1619445, 1619868, 1619596, '0624047', 1619445, 1619868, 2619596, 1619868, 1619596]))

mp()

toc.append(time.time())

print(toc[1]-toc[0])







#загрузка данных в список
from openpyxl import load_workbook
import os

wb = load_workbook(os.path.abspath("d:\OceanWork\PhoenixCo parsing\Terminals.xlsx"), read_only=True) #open file
sheet_terminals = wb['terminals'] #select work sheet

a = 4500
d = sheet_terminals.max_row
c = []
while a <= d:
    c.append(sheet_terminals['G' + str(a)].value)
    a += 1
    print(c)