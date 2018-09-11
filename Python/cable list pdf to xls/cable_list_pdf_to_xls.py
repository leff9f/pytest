import slate3k
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, NamedStyle, Font

print("please input file name without .pdf:")
file_name = input()
# file_name = '1312_06'

# check file exists
while not os.path.isfile(file_name+'.pdf'):
        print("file doesn't exist, please re-input file name, or type exit for exit:")
        file_name = input()
        if file_name == "exit":
            exit(0)

# open pdf file with prop: read boolean
with open(file_name+'.pdf', 'rb') as f:
    doc = slate3k.PDF(f)
# print(doc)

# open txt file for writing and write text from pdf
with open('1312.txt', 'w', encoding='utf-8') as fg:
    fg.write(str(doc))

# open txt file for reading and read text to txt0
with open('1312.txt', 'r', encoding='utf-8', errors='ignore') as file:
    txt = file.readline()

# delete txt temporary file
os.remove('1312.txt')

# text clearing
txt_pr = str(txt).split('\\n')
txt_cleared = []
for a in txt_pr:
    if a != '':
        txt_cleared.append(a)

# cells style
main_style = NamedStyle(name='main_style')
minor_style = NamedStyle(name='minor_style')

thin = Side(border_style='thin', color='000000')
main_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
minor_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
main_style.fill = PatternFill('solid', fgColor='6fb1bd')
minor_style.fill = PatternFill('solid', fgColor='b3dee7')
main_style.font = Font(size=12, name='Calibri')
minor_style.font = Font(size=12, name='Calibri')

# open workbook, rename title
wb = Workbook()
ws = wb.active
ws.title = file_name+'_cable_list'
ws.sheet_properties.tabColor = "20FF20"
wb.add_named_style(main_style)
wb.add_named_style(minor_style)

# change width of column
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 40
ws.column_dimensions["E"].width = 20


# insert style in hat
j = 1
while j != 6:
    ws.cell(row=1, column=j).style = main_style
    j += 1

# hat
a1 = ws.cell(row=1, column=1, value='Кабель')
a2 = ws.cell(row=1, column=2, value='Откуда')
a3 = ws.cell(row=1, column=3, value='Куда')
a4 = ws.cell(row=1, column=4, value='Марка кабеля')
a5 = ws.cell(row=1, column=5, value='Длина')

# import data in excel
txt_sum_nc = txt_cleared[1:]
txt_sum = txt_sum_nc[:len(txt_sum_nc)-1]

j = 2
k = 0
m = 0

for i in txt_sum:
    if m % 5 == 0:
        j += 1
        k = 0
    k += 1
    m += 1
    if k == 1:
        ws.cell(row=int(j), column=int(4), value=txt_sum[m-1])
    if k == 2:
        ws.cell(row=int(j), column=int(5), value=txt_sum[m-1])
    if k == 3:
        ws.cell(row=int(j), column=int(2), value=txt_sum[m-1])
    if k == 4:
        ws.cell(row=int(j), column=int(3), value=txt_sum[m-1])
    if k == 5:
        ws.cell(row=int(j), column=int(1), value=txt_sum[m-1])
    ws.cell(row=int(j), column=int(k)).style = minor_style
    print(j)

# save excel file
wb.save(file_name+'_cable_list.xlsx')

#print(len(txt_pr))
#print(txt_sum)



